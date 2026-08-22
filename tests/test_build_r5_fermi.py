"""R5 build — tools/fermi.py: decomposition + uncertainty propagation.

Hand-derived reference vectors:
  1. Deterministic factors (sigma=0 lognormal) multiply exactly:
     3.0 * 4.0 * 0.5 = 6.0; with zero spread the MC band collapses to a point.
  2. Lognormal multiplication: E[X*Y] = E[X]*E[Y] for independent factors.
     median 100 (sigma 0.2) * median 2 (sigma 0.3): E = 100*2*exp((0.2^2+0.3^2)/2)
     = 200 * exp(0.065) = 213.43...
  3. Independent normals added: std of sum = sqrt(3^2 + 4^2) = 5 (Pythagoras).
  4. Perfectly correlated identical uniforms U(0,1)+U(0,1) with rho->1 gives
     a triangular distribution on (0,2) — mean 1, and the correlated std
     exceeds the independent std (sqrt(1/6) vs sqrt(2/12)=sqrt(1/6)... equal
     for rho=0? No: independent sum std = sqrt(2/12)=0.4082; rho=0.99 pushes
     toward std of 2U = 2/sqrt(12) = 0.5774).
"""
import math
import os
import sys

import numpy as np
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tools.fermi import (
    Correlation,
    Factor,
    FermiResult,
    _bsm_ppf,
    emit_workbook,
    propagate,
)
from tools.artifacts import ArtifactStore


class TestPointPropagation:
    def test_deterministic_product(self):
        r = propagate(
            "test",
            [Factor("a", "lognormal", {"median": 3.0, "sigma": 1e-9}),
             Factor("b", "lognormal", {"median": 4.0, "sigma": 1e-9}),
             Factor("c", "lognormal", {"median": 0.5, "sigma": 1e-9})],
            n_samples=500,
        )
        assert r.mean == pytest.approx(6.0, rel=1e-6)
        assert r.p95 - r.p05 < 1e-6

    def test_lognormal_mean_formula(self):
        r = propagate(
            "rev",
            [Factor("x", "lognormal", {"median": 100.0, "sigma": 0.2}),
             Factor("y", "lognormal", {"median": 2.0, "sigma": 0.3})],
            n_samples=200_000, seed=7,
        )
        expected = 200.0 * math.exp((0.2 ** 2 + 0.3 ** 2) / 2)
        assert r.mean == pytest.approx(expected, rel=0.005)

    def test_pythagoras_independent_addition(self):
        r = propagate(
            "sum",
            [Factor("a", "normal", {"mean": 10.0, "std": 3.0}, combine="add"),
             Factor("b", "normal", {"mean": 20.0, "std": 4.0}, combine="add")],
            n_samples=200_000, seed=3,
        )
        assert r.mean == pytest.approx(30.0, abs=0.1)
        assert r.std == pytest.approx(5.0, rel=0.01)


class TestCorrelation:
    def test_positive_correlation_inflates_sum_std(self):
        f = [Factor("a", "uniform", {"low": 0.0, "high": 1.0}, combine="add"),
             Factor("b", "uniform", {"low": 0.0, "high": 1.0}, combine="add")]
        indep = propagate("s", f, n_samples=100_000, seed=1)
        corr = propagate("s", f, correlations=[Correlation("a", "b", 0.99)],
                         n_samples=100_000, seed=1)
        assert corr.std > indep.std * 1.3
        assert corr.std < 0.60  # bounded by the rho=1 limit of 0.5774 + slack

    def test_marginals_preserved_under_correlation(self):
        f = [Factor("a", "uniform", {"low": 0.0, "high": 10.0}),
             Factor("b", "uniform", {"low": 100.0, "high": 200.0})]
        r = propagate("m", f, correlations=[Correlation("a", "b", 0.9)],
                      n_samples=50_000, seed=5)
        # permutation-based mixing must not distort the marginals
        assert 4.9 < r.factors[0]["sample_median"] < 5.1
        assert 149.0 < r.factors[1]["sample_median"] < 151.0

    def test_determinism(self):
        f = [Factor("a", "lognormal", {"median": 5.0, "sigma": 0.5}),
             Factor("b", "lognormal", {"median": 2.0, "sigma": 0.8})]
        r1 = propagate("d", f, n_samples=10_000, seed=99)
        r2 = propagate("d", f, n_samples=10_000, seed=99)
        assert r1.mean == r2.mean and r1.p95 == r2.p95


class TestAuditTrail:
    def test_factor_rows_carry_sources(self):
        r = propagate(
            "q", [Factor("units", "lognormal", {"median": 1e6, "sigma": 0.3},
                          source="company 10-K", note="trailing units"),
                  Factor("margin", "triangular",
                         {"low": 0.1, "mode": 0.15, "high": 0.25},
                         source="peer set", note="EBITDA margin")],
            unit="USD",
        )
        row = r.factors[0]
        assert row["source"] == "company 10-K"
        assert row["distribution"] == "lognormal"
        assert "params" in row and "sample_median" in row

    def test_sensitivity_ranks_dominant_factor_first(self):
        # wide-sigma factor should dominate a tight one
        r = propagate(
            "s", [Factor("tight", "lognormal", {"median": 10.0, "sigma": 0.01}),
                  Factor("wide", "lognormal", {"median": 10.0, "sigma": 1.5})],
            n_samples=50_000, seed=11,
        )
        assert r.sensitivity[0]["factor"] == "wide"
        assert abs(r.sensitivity[0]["corr_with_result"]) > 0.9

    def test_validation_errors(self):
        with pytest.raises(ValueError):
            propagate("x", [])
        with pytest.raises(ValueError):
            propagate("x", [Factor("a", "weibull", {})])
        with pytest.raises(ValueError):
            propagate("x", [Factor("a", "normal", {"mean": 0, "std": 1}),
                            Factor("a", "normal", {"mean": 0, "std": 1})])
        with pytest.raises(ValueError):
            propagate("x", [Factor("a", "normal", {"mean": 0, "std": 1}),
                            Factor("b", "normal", {"mean": 0, "std": 1})],
                      correlations=[Correlation("a", "ghost", 0.5)])


class TestInverseNormal:
    def test_bsm_matches_scipy_within_tolerance(self):
        try:
            from scipy.stats import norm
        except ImportError:
            pytest.skip("scipy not installed")
        for u in (0.001, 0.024, 0.1, 0.25, 0.5, 0.75, 0.9, 0.975, 0.999):
            assert _bsm_ppf(u) == pytest.approx(float(norm.ppf(u)), abs=4.5e-4)

    def test_bsm_symmetry(self):
        assert abs(_bsm_ppf(0.3) + _bsm_ppf(0.7)) < 1e-9
        assert _bsm_ppf(0.5) == pytest.approx(0.0, abs=1e-12)


class TestWorkbookEmission:
    def test_live_formula_sheet(self, tmp_path):
        r = propagate(
            "2027 revenue", unit="USD",
            factors=[Factor("price", "lognormal", {"median": 50.0, "sigma": 0.2},
                            source="analyst"),
                     Factor("volume", "lognormal", {"median": 2e6, "sigma": 0.4},
                            note="unit growth")],
            n_samples=5_000, seed=1,
        )
        store = ArtifactStore(root=tmp_path / "art")
        out = emit_workbook(r, store=store)
        assert out["live_formulas"] is True
        wb_bytes = store.get_bytes(out["workbook"].sha256)
        assert wb_bytes[:2] == b"PK"  # real xlsx

        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(wb_bytes))
        # live formula chains Assumptions cells: B2 * B3
        assert wb["ModelLive"]["B1"].value in ("=B2*B3", "=B3*B2")
        # every assumption labelled with distribution + source
        a_sheet = wb["Assumptions"]
        notes = " ".join(str(c.value) for row in a_sheet.iter_rows() for c in row)
        assert "lognormal" in notes and "analyst" in notes

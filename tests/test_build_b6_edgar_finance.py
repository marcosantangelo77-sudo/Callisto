"""B6 tests — offline fixture suite for the EDGAR/finance domain.

No network: every test runs against a synthetic companyfacts payload that
exercises the real mess (retired tags, restatements, mixed fiscal years,
missing concepts, non-USD units). Live SEC verification was done during
the build and is documented in findings/instance_b6.md.
"""
from __future__ import annotations

import asyncio
import io
from typing import Optional

import pytest

from tools.domains.finance.edgar import (
    EdgarClient,
    annual_facts,
    instant_facts,
)
from tools.domains.finance.statements import (
    FinancialStatements,
    assemble_statements,
)
from tools.domains.finance.models import (
    comps_workbook,
    dcf_workbook,
    dcf_sandbox_code,
    proforma_workbook,
)


# ── fixtures ──────────────────────────────────────────────────────────────

def _fact(start, end, val, *, accn="0001-24", form="10-K", filed="2024-08-01",
          fy=2024, fp="FY", unit="USD"):
    f = {"start": start, "end": end, "val": val, "accn": accn,
         "fy": fy, "fp": fp, "form": form, "filed": filed}
    if unit != "USD":
        return {unit: [f]}
    return {"USD": [f]}


def make_facts() -> dict:
    """Synthetic TestCo: June FYE, retired revenue tag, a restatement.

    History:
      - old tag Revenues used FY2018–FY2020 (retired thereafter);
      - current tag RevenueFromContractWithCustomerExcludingAssessedTax
        covers FY2021–FY2023;
      - FY2022 net income was RESTATED in a later filing;
      - dividends are reported in EUR (non-USD unit) to prove unit handling;
      - operating_expenses is missing entirely → explicit gap.
    """
    usgaap = {
        "Revenues": {
            "label": "Revenues",
            "units": _merge(
                _fact("2019-07-01", "2020-06-30", 50e9),
                _fact("2020-07-01", "2021-06-30", 60e9),
            ),
        },
        "RevenueFromContractWithCustomerExcludingAssessedTax": {
            "label": "RevenueFromContractWithCustomerExcludingAssessedTax",
            "units": _merge(
                _fact("2021-07-01", "2022-06-30", 80e9),
                _fact("2022-07-01", "2023-06-30", 100e9),
                # comparative duplicate of FY2022 in next year's 10-K:
                _fact("2021-07-01", "2022-06-30", 80e9, accn="0002-25",
                      filed="2025-08-01", fy=2025),
                # restated FY2022 in an amended filing:
                _fact("2021-07-01", "2022-06-30", 78e9, accn="0003-25",
                      form="10-K/A", filed="2025-09-15", fy=2025),
            ),
        },
        "CostOfRevenue": {
            "label": "CostOfRevenue",
            "units": _merge(
                _fact("2021-07-01", "2022-06-30", 40e9),
                _fact("2022-07-01", "2023-06-30", 45e9),
            ),
        },
        "NetIncomeLoss": {
            "label": "NetIncomeLoss",
            "units": _merge(
                _fact("2021-07-01", "2022-06-30", 12e9, accn="0000-22",
                      filed="2022-08-01", fy=2022),
                _fact("2021-07-01", "2022-06-30", 10.5e9, accn="0004-25",
                      filed="2025-09-15", fy=2025),   # restated DOWN
                _fact("2022-07-01", "2023-06-30", 18e9),
            ),
        },
        "EarningsPerShareDiluted": {
            "label": "EPS diluted",
            "units": {
                "USD/shares": [
                    {"start": "2022-07-01", "end": "2023-06-30", "val": 3.6,
                     "accn": "0001-24", "fy": 2024, "fp": "FY",
                     "form": "10-K", "filed": "2024-08-01"},
                ],
            },
        },
        "Assets": {"units": _merge(_fact(None, "2023-06-30", 300e9))},
        "CashAndCashEquivalentsAtCarryingValue": {
            "units": _merge(_fact(None, "2023-06-30", 25e9))},
        "AssetsCurrent": {"units": _merge(_fact(None, "2023-06-30", 120e9))},
        "LiabilitiesCurrent": {"units": _merge(_fact(None, "2023-06-30", 80e9))},
        "LongTermDebtNoncurrent": {"units": _merge(_fact(None, "2023-06-30", 50e9))},
        "StockholdersEquity": {"units": _merge(_fact(None, "2023-06-30", 180e9))},
        "NetCashProvidedByUsedInOperatingActivities": {
            "units": _merge(
                _fact("2021-07-01", "2022-06-30", 28e9),
                _fact("2022-07-01", "2023-06-30", 33e9),
            ),
        },
        "PaymentsToAcquirePropertyPlantAndEquipment": {
            "units": _merge(
                _fact("2021-07-01", "2022-06-30", 7e9),
                _fact("2022-07-01", "2023-06-30", 9e9),
            ),
        },
        "PaymentsOfDividendsCommonStock": {
            "label": "Dividends paid",
            "units": _merge(
                _fact("2022-07-01", "2023-06-30", 4.0e9, unit="EUR"),
            ),
        },
    }
    return {"cik": 999999, "entityName": "TESTCO INC",
            "facts": {"us-gaap": usgaap}}


def _merge(*unit_dicts) -> dict:
    out: dict[str, list] = {}
    for ud in unit_dicts:
        for u, fs in ud.items():
            out.setdefault(u, []).extend(fs)
    return out


# ── edgar.py fact selection ───────────────────────────────────────────────

class TestFactSelection:
    def test_annual_filters_quarterlies(self):
        facts = {"facts": {"us-gaap": {"X": {"units": _merge(
            _fact("2023-01-01", "2023-03-31", 1.0),
            _fact("2022-07-01", "2023-06-30", 4.0),
        )}}}}
        got = annual_facts(facts, "X")
        assert len(got) == 1 and got[0]["val"] == 4.0

    def test_latest_filing_wins_restatement(self):
        facts = make_facts()
        ni = {f["start"]: f["val"] for f in annual_facts(facts, "NetIncomeLoss")}
        assert ni["2021-07-01"] == 10.5e9  # restated value, not original 12e9

    def test_retired_tag_facts_still_retrievable(self):
        facts = make_facts()
        old = annual_facts(facts, "Revenues")
        assert len(old) == 2 and old[-1]["end"] == "2021-06-30"

    def test_instant_ignores_duration(self):
        facts = {"facts": {"us-gaap": {"Y": {"units": _merge(
            _fact(None, "2023-06-30", 5.0),
            _fact("2022-07-01", "2023-06-30", 99.0),
        )}}}}
        got = instant_facts(facts, "Y")
        assert len(got) == 1 and got[0]["val"] == 5.0


# ── statements.py assembly ────────────────────────────────────────────────

class TestAssembly:
    def _stmt(self, n=3) -> FinancialStatements:
        s = assemble_statements(make_facts(), n_periods=n)
        s.ticker = "TST"
        return s

    def test_anchor_picks_most_recent_tag(self):
        s = self._stmt()
        # must NOT anchor on retired Revenues (ends 2021) — periods are recent.
        # n_periods=3 asks for 3 periods but the current tag only reaches back
        # two, so FY2021 is simply absent (the retired tag's facts are not
        # silently spliced in — that would mix tag bases across a column).
        assert list(s.periods) == ["FY2022", "FY2023"]
        assert s.used_tags["revenue"].startswith("RevenueFromContract")

    def test_restatement_flagged_not_hidden(self):
        s = self._stmt()
        ni = [l for l in s.income if l.label == "net_income"
              and l.period == "FY2022"]
        assert ni and ni[0].restated is True
        assert ni[0].value == 10.5e9

    def test_missing_concept_is_gap_not_zero(self):
        s = self._stmt()
        opex = [l for l in s.income if l.label == "operating_expenses"]
        assert all(l.value is None for l in opex)
        assert any("operating_expenses" in g for g in s.gaps)

    def test_non_usd_unit_carried(self):
        # A fact reported in another unit (EUR dividends) must surface with
        # its own unit — never silently converted or relabelled USD, which
        # would fabricate a number the filing does not contain.
        s = self._stmt()
        div = [l for l in s.cash_flow if l.label == "dividends_paid"]
        eur = [l for l in div if l.unit == "EUR" and l.value == 4.0e9]
        assert eur, "EUR-denominated dividend must be surfaced with its unit"
        usd_cells = [l for l in div if l.unit == "USD"]
        assert all(l.value is None for l in usd_cells if not l.restated)

    def test_derived_lines_marked(self):
        s = self._stmt()
        gp = {l.period: l for l in s.derived if l.label == "gross_profit"}
        fcf = [l for l in s.derived if l.label == "free_cash_flow"]
        wc = [l for l in s.derived if l.label == "working_capital"]
        assert gp and all(l.derived and "revenue - cost_of_revenue" in l.derivation
                          for l in gp.values())
        assert gp["FY2023"].value == 55e9  # 100 - 45
        assert fcf and fcf[-1].value == 24e9
        assert wc and wc[-1].value == 40e9

    def test_every_value_traceable_to_filing(self):
        s = self._stmt()
        reported = [l for l in s.income + s.balance_sheet + s.cash_flow
                    if l.value is not None and not l.derived]
        assert reported
        for lv in reported:
            assert lv.tag and lv.accn and lv.form and lv.filed

    def test_limitations_always_declared(self):
        s = self._stmt()
        assert any("footnote" in lim.lower() for lim in s.limitations)

    def test_no_revenue_raises(self):
        with pytest.raises(ValueError):
            assemble_statements({"facts": {"us-gaap": {}}})


# ── models.py templates ───────────────────────────────────────────────────

class TestModels:
    def _stmt(self) -> FinancialStatements:
        return assemble_statements(make_facts(), n_periods=3)

    def test_dcf_spec_shape_and_live_cells(self):
        spec, notes = dcf_workbook(self._stmt())
        assert notes["wacc_placeholder"] is True  # judgment flagged as such
        cells = {m["cell"]: m["formula"] for m in spec["model"]}
        # formulas reference Assumptions — live chain, not baked values
        assert cells["B3"].startswith("Assumptions!$B$2*(1+")
        assert "(Assumptions!$B$7-Assumptions!$B$8)" in cells["B9"]
        assert any("SUM(B6:F6)" == f for f in cells.values())
        srcs = {a["name"]: a["source"] for a in spec["assumptions"]}
        assert "XBRL" in srcs["base_revenue"]
        assert "REVIEW" in srcs["wacc"]

    def test_dcf_sandbox_matches_hand_math(self):
        code = dcf_sandbox_code(110e9, 0.10, 0.227, 0.10, 0.02, 5)
        from tools.sandbox import run_python

        res = run_python(code)
        assert res.status == "ok" and res.return_value
        ev = res.return_value["enterprise_value"]
        # independent hand computation of the same closed-form chain
        rev, pv, fcfs = None, 0.0, []
        r = 110e9 * 1.10
        m = 0.227
        for yr in range(1, 6):
            if yr > 1:
                g = 0.10 + (0.03 - 0.10) * ((yr - 1) / 4)
                r = r * (1 + g)
            fcfs.append(r * m)
            pv += fcfs[-1] / 1.1 ** yr
        tv = fcfs[-1] * 1.02 / 0.08
        assert abs(ev - (pv + tv / 1.1 ** 5)) < 1.0

    def test_dcf_analyst_inputs_override_seeds(self):
        spec, notes = dcf_workbook(
            self._stmt(),
            analyst_inputs={"wacc": 0.085, "terminal_growth": 0.025,
                            "diluted_shares": 1e9})
        srcs = {a["name"]: a["source"] for a in spec["assumptions"]}
        vals = {a["name"]: a["value"] for a in spec["assumptions"]}
        assert srcs["wacc"] == "analyst input"
        assert vals["wacc"] == 0.085 and notes["wacc_placeholder"] is False

    def test_dcf_workbook_emits_real_xlsx_with_formulas(self):
        from tools.charts import build_workbook

        xlsx = build_workbook(dcf_workbook(self._stmt())[0])
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["ModelLive"]
        assert str(ws["B13"].value).startswith("=IF(Assumptions!$B$10>0")

    def test_proforma_articulates(self):
        from tools.charts import build_workbook

        spec, _notes = proforma_workbook(self._stmt())
        cells = {m["cell"]: m["formula"] for m in spec["model"]}
        # CFO = NI − ΔNWC; net cash flow sums the pieces; plug accumulates
        assert cells["B7"].replace(" ", "") == "B3-B6"
        assert cells["B10"].replace(" ", "") == "B7+B8-B9"
        assert cells["C11"].replace(" ", "") == "B11+C10"
        build_workbook(spec)  # openpyxl accepts every formula cell

    def test_comps_multiples_are_formulas(self):
        peers = [
            {"name": "A", "price": 100, "shares": 1e6, "eps": 5.0,
             "bvps": 40.0, "ebitda": 900e6, "revenue": 2e9,
             "debt": 100e6, "cash": 50e6},
            {"name": "B", "price": 55, "shares": 2e6, "eps": 4.0, "bvps": 30.0},
        ]
        spec, _notes = comps_workbook(peers)
        cells = {m["cell"]: m["formula"] for m in spec["model"]}
        assert cells["K2"] == 'IF(D2>0,J2/(D2*C2),"nm")'   # P/E blank-safe
        assert cells["N2"] == 'IF(F2>0,M2/F2,"nm")'
        assert cells["M2"] == "J2+SUM(H2:I2)"              # EV = cap + net debt
        medians = [f for f in cells.values() if f.startswith("MEDIAN")]
        assert len(medians) == 3
        from tools.charts import build_workbook

        build_workbook(spec)

    def test_comps_requires_peers(self):
        with pytest.raises(ValueError):
            comps_workbook([])


# ── plugin registration + dispatch ────────────────────────────────────────

class TestPlugin:
    def test_registers_and_serves_financial(self):
        from tools.domain_registry import ToolRegistry
        from tools.domains.finance.plugin import register_if_available
        from agp import Domain

        reg = ToolRegistry()
        assert register_if_available(reg) is True
        names = reg.tool_names_for(Domain.FINANCIAL, "")
        assert {"edgar_get_statements", "edgar_build_model"} <= names

    def test_does_not_serve_sports_query(self):
        from tools.domain_registry import ToolRegistry
        from tools.domains.finance.plugin import build_finance_plugin
        from agp import Domain

        reg = ToolRegistry()
        reg.register(build_finance_plugin())
        # declared-domain plugins are not pulled in by keywords alone
        names = reg.tool_names_for(Domain.TECHNICAL, "dcf valuation of NVDA")
        assert not names

    def test_dispatch_error_is_structured_not_raised(self):
        from tools.domain_registry import ToolRegistry
        from tools.domains.finance.plugin import register_if_available

        reg = ToolRegistry()
        register_if_available(reg)
        handled, res = asyncio.run(reg.dispatch(
            "edgar_get_statements", {"ticker": ""}))
        assert handled and res["ok"] is False and "error" in res

    def test_dispatcher_offline_failure_reports_cleanly(self, monkeypatch):
        """With network broken, tool returns ok=False instead of raising."""
        from tools.domain_registry import ToolRegistry
        from tools.domains.finance import plugin as fp
        from tools.domains.finance.edgar import EdgarError

        reg = ToolRegistry()
        reg.register(fp.build_finance_plugin())

        class BrokenClient:
            def facts_for_ticker(self, t):
                raise EdgarError("offline")

        monkeypatch.setattr(fp, "_get_client", lambda: BrokenClient())
        handled, res = asyncio.run(reg.dispatch(
            "edgar_build_model",
            {"template": "dcf", "ticker": "XYZ"}))
        assert handled and res["ok"] is False


# ── provenance wiring ─────────────────────────────────────────────────────

class TestProvenanceWiring:
    def test_client_records_primary_observation_in_ledger(self):
        from agp.provenance import ProvenanceLedger

        class FakeClient(EdgarClient):
            def __init__(self):
                super().__init__(ledger=ProvenanceLedger(), min_interval_s=0)

        c = FakeClient.__new__(FakeClient)
        EdgarClient.__init__(c, ledger=ProvenanceLedger(), min_interval_s=0)
        body = '{"cik": 123}'
        c._record("https://example.gov/x.json", 200, body, 0.01)
        assert c.ledger.has_observation(body)
        assert c.ledger.is_primary_bytes(body)
        rec = c._last_record
        assert rec.content_sha256 and rec.status == 200

    def test_ledger_none_is_tolerated(self):
        c = EdgarClient(ledger=None, min_interval_s=0)
        rec = c._record("u", 200, "{}", 0.0)  # no ledger: fetch still records
        assert rec.size_bytes == 2

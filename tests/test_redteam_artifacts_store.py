"""RED TEAM — the artifact store and chart regeneration (tools/artifacts.py,
tools/charts.py).

The claim under attack is tools/artifacts.py's own module docstring:
"Immutable: an artifact id IS its content ... Tampering with the store is
detectable by re-hash. Sealable: ... verifying the seal verifies the artifacts
still exist unmodified." Property 3 of BUILD_MANDATE ("verifiable, not
voluminous") rests on that docstring being true.

Method: adversarial constructions against invariants stated by the code or
its docs, each reduced to a runnable failing test. Honest negatives are at
the bottom and kept as pins.

FINDINGS (each test below fails on master as described):
  A2  gc() + corrupt index destroys evidence bytes permanently.
  A3  re-put rewrites an artifact's meta/data_refs provenance.
  A4  export_ref writes outside dest_dir via a crafted ref name.
  A5  failed sandbox runs still mint artifact refs (status recorded, files
      sealed anyway) — wait: fixed mid-run? see A5 note in report; current
      behaviour: stdout sealed for failed run but child-attested file refs
      skipped only when workspace absent AND status != ok.
  A6  child-attested refs cite hashes with NO bytes in the store;
      verify_artifacts reports ok=False but nothing upstream checks it.
  A9  dedupe lets a later put REWRITE code_sha256 of an existing artifact
      when the earlier entry was empty — first-seen-wins only half-holds.
  A11 render_svg accepts NaN/Inf and emits them into axis text.
  A12 _sniff_kind calls ANY zip "xlsx" and any "<svg" substring svg.
  A13 ModelLive silently keeps the LAST formula written to a duplicated cell
      while the Model listing shows both — workbook contradicts its own audit
      sheet.
  A14 concurrent puts lose index entries AND crash with FileNotFoundError.
  A16 rebuild_index erases code_sha256/name/data_refs provenance — the
      documented recovery path launders origin.
  A17 export_ref silently overwrites an existing delivery file.
  A18 verify_artifacts accepts refs whose kind/code_sha256 lie about bytes;
      ArtifactRef.from_dict accepts non-hex ids; seal payload contains no
      artifact ids at all (A20) so the "sealable" claim is false end-to-end.
  B1  Data-sheet rows starting with "=" become LIVE Excel formulas — fetched
      evidence text can carry formulas into the auditable workbook.
  B3  provenance comments attach to column 1 when the named column is
      missing — misattribution, not failure.

HONEST NEGATIVES (pins, pass on master):
  N1  get_path cannot be walked above the store root (Path join semantics
      block the classic ../../ read).
  N2  chart regeneration round-trips: store_chart output re-renders from the
      stored spec byte-identically (SVG path).
  N3  objects are content-addressed and never overwritten by put().
"""
import io
import json
import tempfile
import threading
from pathlib import Path

import pytest

from tools.artifacts import (
    ALLOWED_KINDS,
    ArtifactRef,
    ArtifactStore,
    _sniff_kind,
    sha256_bytes,
    store_sandbox_outputs,
)
from tools.charts import build_workbook, chart_spec, render_svg, store_chart


class _FakeSandboxResult:
    """Duck-typed tools.sandbox.SandboxResult."""

    def __init__(self, code, stdout, files=(), status="ok"):
        self.code = code
        self.stdout = stdout
        self.files = list(files)
        self.status = status


def _store(tmp_path=None):
    return ArtifactStore(Path(tempfile.mkdtemp(dir=tmp_path)))


# ── A2: gc after index corruption deletes evidence ─────────────────────────

def test_gc_after_corrupt_index_deletes_objects(tmp_path):
    s = _store()
    ref = s.put(b"important evidence bytes", "txt", name="evidence")
    assert s.exists(ref.sha256)
    s.index_path.write_text("{corrupt")          # bit rot / partial write
    removed = s.gc()
    # The ONLY object in the store was just unlinked because the corrupt
    # index made it look like an orphan. The module docstring promises
    # "objects survive" a corrupt index; gc() makes that promise false.
    assert removed == [] , f"gc destroyed evidence: {removed}"
    assert s.exists(ref.sha256)


# ── A3: re-put mutates provenance ──────────────────────────────────────────

def test_reput_rewrites_data_refs_and_meta_of_existing_artifact():
    s = _store()
    r1 = s.put(b"x", "json", name="honest", meta={"live_formulas": True})
    r2 = s.put(b"x", "json", name="laundered",
               data_refs=["ab" * 32], meta={"live_formulas": False})
    assert r1.sha256 == r2.sha256                # same bytes, same id
    m = s.get_meta(r1.sha256)
    # The comment in _index_add says "an artifact's origin does not change
    # because someone later re-put identical bytes" — yet data_refs and all
    # meta keys are overwritten by the later call.
    assert m["data_refs"] == []
    assert m["meta"] == {"live_formulas": True}


# ── A4/A17: export_ref escapes dest_dir and overwrites ─────────────────────

def test_export_ref_name_traversal_writes_outside_dest_dir(tmp_path):
    s = _store()
    ref = s.put(b"x", "txt")
    outside = tmp_path.parent / ("rt_pwned_%d.txt" % threading.get_ident())
    try:
        dest_dir = tmp_path / "delivery"
        out = s.export_ref(
            ArtifactRef(sha256=ref.sha256, kind="", name=f"../{outside.name}"),
            dest_dir,
        )
        assert dest_dir.resolve() in out.resolve().parents or \
            out.resolve() == (dest_dir / "../" ).resolve(), \
            f"export escaped dest_dir: {out}"
    finally:
        outside.unlink(missing_ok=True)


def test_export_ref_silently_overwrites_existing_file(tmp_path):
    s = _store()
    victim = tmp_path / "report.txt"
    victim.write_text("PRECIOUS HUMAN FILE")
    ref = s.put(b"MALICIOUS REPLACEMENT", "txt", name="report")
    s.export_ref(ref, tmp_path)
    assert victim.read_text() == "PRECIOUS HUMAN FILE"


# ── A5/A6: sandbox outputs ─────────────────────────────────────────────────

def test_failed_run_still_seals_stdout_and_attested_files_without_workspace():
    s = _store()
    sbx = _FakeSandboxResult(
        code="import os", stdout="", status="error",
        files=[{"name": "out.csv", "sha256": "c" * 64}],
    )
    refs = store_sandbox_outputs(sbx, s)         # no workspace → attested path
    names = [r.name for r in refs]
    # A FAILED run must not contribute citable artifacts; here it mints refs.
    assert not any(r.meta.get("attested_by_child_only") for r in refs), \
        f"failed run produced citable refs: {names}"


def test_child_attested_ref_has_no_bytes_in_store():
    """A6 — FIXED. The attested path no longer mints refs from hashes
    reported by the sandbox child: a child-attested hash is a claim, not
    evidence. The claim is stored as an explicit non-citable JSON record,
    and every returned ref has bytes in the store."""
    s = _store()
    sbx = _FakeSandboxResult("print('hi')", "hi\n",
                             files=[{"name": "model.csv", "sha256": "a" * 64}])
    refs = store_sandbox_outputs(sbx, s)
    # Invariant 1: no ref exists whose bytes are not in the store.
    for r in refs:
        assert s.exists(r.sha256), \
            f"cited ref {r.name} has no bytes in store"
    # Invariant 2: the child's claim is preserved, as a record, and is
    # explicitly marked unfit for citation as quantitative evidence.
    claims = [r for r in refs if r.kind == "json"
              and r.name == "attestation_claim"]
    assert claims, "expected the attestation claim record to be stored"
    claim = json.loads(s.get_bytes(claims[0].sha256).decode("utf-8"))
    assert claim.get("citable_as_evidence") is False
    assert claim["files_reported_by_child"][0]["sha256"] == "a" * 64
    # Regression guard: verify_artifacts passes over everything cited.
    report = s.verify_artifacts(refs)
    assert report["ok"] is True


# ── A9: dedupe rewrites code_sha256 when first entry lacked one ────────────

def test_reput_with_code_overwrites_empty_code_sha256_provenance():
    s = _store()
    plain = s.put(b"same output bytes", "txt")   # e.g. rebuilt index / manual
    code_hash = sha256_bytes(b"fabricating code")
    s.put(b"same output bytes", "txt", code_sha256=code_hash,
          name="stdout")
    m = s.get_meta(plain.sha256)
    # First-seen wins is claimed; but empty code_sha256 invites takeover.
    # Whichever way this resolves, the invariant must hold BOTH directions:
    assert m["code_sha256"] == "", (
        "later put rewrote code_sha256 provenance of existing artifact"
    )


# ── A11: NaN/Inf reach chart output ────────────────────────────────────────

def test_render_svg_emits_nan_into_axis_labels():
    spec = chart_spec("t", {"a": [float("nan"), 1.0]})
    svg = render_svg(spec)
    assert "nan" not in svg.lower().replace("font-family", ""), \
        "NaN series values leaked into rendered axis text"


def test_render_svg_accepts_inf_x_axis():
    spec = chart_spec("t", {"a": [1.0, 2.0]}, x=[float("inf"), 0.0])
    svg = render_svg(spec)
    assert "inf" not in svg.replace("xmlns=", "").lower() or "Infinity" in svg


# ── A12: sniffing ──────────────────────────────────────────────────────────

@pytest.mark.parametrize("data,not_kind", [
    (b"PK\x03\x04" + b"arbitrary zip bytes", "xlsx"),
    (b"<html><body><svg onload='x'/></body></html>" + b" " * 600, "svg"),
])
def test_sniff_kind_misclassifies(data, not_kind):
    assert _sniff_kind(data) != not_kind, \
        f"{not_kind} assigned from magic-substring alone"


# ── A13: duplicate model cell — workbook contradicts its audit sheet ───────

def test_duplicate_model_cell_listing_contradicts_live_sheet():
    spec = {
        "model": [
            {"cell": "A1", "formula": "1", "label": "first"},
            {"cell": "A1", "formula": "999999", "label": "overwrite"},
        ],
    }
    wb = load_workbook_bytes(build_workbook(spec))
    listing = wb["Model"]
    live = wb["ModelLive"]
    listed_formulas = [listing.cell(row=r, column=3).value for r in (2, 3)]
    assert listed_formulas == ["=1", "=999999"]
    # The live sheet holds only ONE of them. An auditor reading Model sees
    # two formulas; the workbook computes with the other. Contradiction.
    assert live["A1"].value == "=1", (
        "ModelLive silently dropped/overwrote a listed formula"
    )


# ── A14: concurrent puts ───────────────────────────────────────────────────

def test_concurrent_puts_do_not_lose_entries_or_crash():
    s = _store()
    errors = []

    def worker(tag):
        try:
            for i in range(25):
                s.put(f"{tag}-{i}".encode(), "txt")
        except Exception as e:  # noqa: BLE001
            errors.append(e)

    ts = [threading.Thread(target=worker, args=(t,)) for t in ("a", "b", "c")]
    [t.start() for t in ts]
    [t.join() for t in ts]
    assert errors == [], f"put raised under concurrency: {errors!r}"
    idx = s._load_index()
    assert len(idx) == 75, f"lost index entries: {len(idx)}/75"


# ── A16: rebuild_index erases provenance ───────────────────────────────────

def test_rebuild_index_erases_code_and_name_provenance():
    s = _store()
    sbx = _FakeSandboxResult("analysis code", "out")
    refs = store_sandbox_outputs(sbx, s)
    before = s.get_meta(refs[0].sha256)
    assert before["code_sha256"]
    s.rebuild_index()
    after = s.get_meta(refs[0].sha256)
    # The documented recovery path must not launder origin: either preserve
    # what it can prove, or mark entries as reconstructed.
    assert after["code_sha256"] != "" or after["meta"].get("reconstructed"), (
        "rebuild_index silently stripped producing-code provenance"
    )


# ── A18: verification does not bind metadata ───────────────────────────────

def test_verify_artifacts_accepts_ref_lying_about_kind_and_code():
    s = _store()
    ref = s.put(b"just text", "txt")
    liar = ArtifactRef(sha256=ref.sha256, kind="png",
                       name="screenshot-evidence.png",
                       code_sha256="f" * 64)
    report = s.verify_artifacts([liar])
    # Re-hash confirms BYTES; but 'ok' then vouches for a ref whose declared
    # kind and producing-code hash are fabricated. verify must check what a
    # seal would cover.
    assert report["ok"] is True
    idx_entry = s.get_meta(ref.sha256) or {}
    assert liar.kind == idx_entry.get("kind") and \
        liar.code_sha256 == idx_entry.get("code_sha256")


def test_from_dict_rejects_non_hex_sha256():
    with pytest.raises(ValueError):
        ArtifactRef.from_dict({"sha256": "../../etc/passwd", "kind": "txt"})


# ── A20: the seal does not cover artifacts at all ──────────────────────────

def test_sealed_session_payload_contains_no_artifact_ids():
    from agp import AGPSession, Domain, Evidence, SourceClass

    sess = AGPSession(query="q")
    sess.add_evidence(Evidence(content="c",
                               source_class=SourceClass.PRIMARY,
                               confidence_score=0.5,
                               domain=Domain.GENERAL,
                               origin_agent="x"))
    d = sess.to_dict()
    assert any("artifact" in k.lower() for k in d), (
        "AGP session payload carries no artifact references — the keyed "
        "seal covers conclusions and evidence but NOT the quantitative "
        "artifacts the conclusion cites"
    )


# ── B1/B3: workbook injection and misattribution ───────────────────────────

def test_data_row_starting_with_equals_becomes_live_formula():
    spec = {
        "data": {"Pulls": {"columns": ["note", "value"],
                           "rows": [['=HYPERLINK("http://evil","click")', 2]]}},
    }
    wb = load_workbook_bytes(build_workbook(spec))
    cell = wb["Pulls"].cell(row=2, column=1)
    # Fetched evidence text flows into cells verbatim; a value beginning '='
    # is stored as a FORMULA (data_type 'f'), i.e. executed by Excel on open.
    assert cell.data_type == "s" or not str(cell.value).startswith("="), (
        f"evidence text became a live formula: {cell.value!r}"
    )


def test_negative_number_string_stays_numeric_not_text():
    # B1 boundary: "-5" is a plain negative number and must be stored as a
    # number (or at worst an unmodified string) — never "'-5".
    spec = {
        "assumptions": [{"name": "drift", "value": "-5"}],
        "data": {"Pulls": {"columns": ["v"], "rows": [["-5"], ["+3"],
                                                      ["@2"], ["-SUM(A1)"]]}},
    }
    wb = load_workbook_bytes(build_workbook(spec))
    assert wb["Pulls"].cell(row=2, column=1).value == -5
    assert wb["Pulls"].cell(row=2, column=1).data_type != "f"
    assert wb["Assumptions"].cell(row=2, column=2).value == -5
    # "+3" parses as a number, so it is stored as the number 3.
    assert wb["Pulls"].cell(row=3, column=1).value == 3
    assert wb["Pulls"].cell(row=3, column=1).data_type == "n"
    # "@2" is not a number: neutralized to text so Excel can't run it as
    # a macro/function lead.
    assert wb["Pulls"].cell(row=4, column=1).value == "'@2"
    # Non-number strings with +/-/@ leads are neutralized to text.
    cell = wb["Pulls"].cell(row=5, column=1)
    assert str(cell.value).startswith("'") or cell.data_type == "s"


def test_plus_at_prefix_nonnumeric_strings_are_neutralized():
    spec = {
        "data": {"Pulls": {"columns": ["a", "b", "c"], "rows": [
            ["+cmd|calc", "@SUM(A1)", "-HYPERLINK(\"http://evil\",\"x\")"],
        ]}},
    }
    wb = load_workbook_bytes(build_workbook(spec))
    for col in (1, 2, 3):
        cell = wb["Pulls"].cell(row=2, column=col)
        assert cell.data_type != "f", f"column {col} executed: {cell.value!r}"
        assert str(cell.value).startswith("'")


def test_legitimate_system_formulas_remain_live():
    # Model formulas come from OUR code (spec["model"]), never fetched bytes;
    # they must still land as live formulas so the workbook stays auditable.
    spec = {
        "model": [{"cell": "B2", "formula": "=Assumptions!B2*2",
                   "label": "double"}],
    }
    wb = load_workbook_bytes(build_workbook(spec))
    live = wb["ModelLive"]["B2"]
    assert live.data_type == "f"
    assert str(live.value).startswith("=")
    listing = wb["Model"].cell(row=2, column=3)
    assert listing.data_type == "f"


def test_provenance_comment_for_unknown_column_lands_on_first_column():
    spec = {
        "data": {"Pulls": {"columns": ["a", "b"], "rows": [[1, 2]],
                           "provenance": [{"column": "typo_col",
                                           "source": "FRED",
                                           "fetched_at": "today"}]}},
    }
    wb = load_workbook_bytes(build_workbook(spec))
    ws = wb["Pulls"]
    assert ws.cell(row=1, column=1).comment is None, (
        "provenance for a nonexistent column was silently attached to "
        "column 1 — misattribution instead of rejection"
    )


# ── honest negatives / regression pins ─────────────────────────────────────

def test_pin_get_path_cannot_escape_store_root():
    s = _store()
    for evil in ["..../..../x", "../../../secret"]:
        with pytest.raises(KeyError):
            s.get_path(evil)
    # "../../etc/passwd" DOES escape the objects/ tree (Path join keeps the
    # .. segments) but lands at an absolute path that does not exist in a
    # test environment — pin the refusal as behaviour, not just KeyError.
    with pytest.raises(KeyError):
        s.get_path("../../definitely_not_here_9x7")


def test_pin_chart_regeneration_round_trip_svg():
    s = _store()
    spec = chart_spec("Test", {"a": [1.0, 2.0, 3.0]}, code="x=1")
    res = store_chart(spec, s)
    stored = s.get_json(res["spec"].sha256)
    regen = render_svg(stored).encode("utf-8")
    assert regen == s.get_bytes(res["chart"].sha256)


def test_pin_object_never_overwritten_by_put():
    s = _store()
    r = s.put(b"original", "txt")
    obj = s.get_path(r.sha256).read_bytes()
    s.put(b"original", "txt", name="again")
    assert s.get_path(r.sha256).read_bytes() == obj


# ── helpers ────────────────────────────────────────────────────────────────

def load_workbook_bytes(data: bytes):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(data))

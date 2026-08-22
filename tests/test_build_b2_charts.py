"""B2 build pass — charts + live-formula spreadsheets (tools/charts.py)."""
import io
import json
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tools.artifacts import ArtifactStore
from tools.charts import (
    build_workbook,
    build_workbook_csv_fallback,
    chart_spec,
    render_svg,
    store_chart,
    store_workbook,
)


@pytest.fixture()
def store(tmp_path):
    return ArtifactStore(root=tmp_path / "artifacts")


class TestChartSpec:
    def test_spec_carries_code_hash(self):
        s = chart_spec("t", {"a": [1, 2]}, code="x=1")
        assert s["code_sha256"]

    def test_length_mismatch_rejected(self):
        with pytest.raises(ValueError):
            chart_spec("t", {"a": [1, 2], "b": [1, 2, 3]})

    def test_x_length_mismatch_rejected(self):
        with pytest.raises(ValueError):
            chart_spec("t", {"a": [1, 2]}, x=[1])


class TestSvgRenderer:
    def test_deterministic(self):
        s = chart_spec("t", {"a": [1, 2, 3], "b": [2, 1, 4]}, x=[10, 20, 30])
        assert render_svg(s) == render_svg(s)

    def test_contains_series_and_escapes(self):
        s = chart_spec("a<b & c", {"s1": [1, 2]})
        svg = render_svg(s)
        assert svg.startswith("<svg")
        assert "a&lt;b &amp; c" in svg
        assert svg.count("<polyline") == 1

    def test_two_series_two_polylines(self):
        s = chart_spec("t", {"x": [1, 2], "y": [3, 4]})
        assert render_svg(s).count("<polyline") == 2


class TestStoreChart:
    def test_chart_and_spec_linked(self, store):
        out = store_chart(
            chart_spec("supply", {"s": [1, 2, 3]}, code="gen()"), store
        )
        chart_meta = store.get_meta(out["chart"].sha256)
        spec = store.get_json(out["spec"].sha256)
        assert out["spec"].sha256 in chart_meta["data_refs"]
        assert spec["code"] == "gen()"
        assert out["renderer"] in ("matplotlib", "svg")
        assert out["chart"].kind in ("png", "svg")

    def test_chart_bytes_deterministic_svg(self, store):
        spec = chart_spec("t", {"a": [1, 2]})
        r1 = store_chart(spec, store)
        r2 = store_chart(spec, store)
        assert r1["chart"].sha256 == r2["chart"].sha256


class TestWorkbook:
    def _spec(self):
        return {
            "assumptions": [
                {"name": "ratio", "value": 58.0, "unit": "x",
                 "source": "derived", "note": "post-halving"},
            ],
            "data": {
                "Prices": {
                    "columns": ["year", "px"],
                    "rows": [[2024, 65000]],
                    "provenance": [{"column": "px", "source": "src",
                                    "fetched_at": "2026-08-22"}],
                }
            },
            "model": [
                {"cell": "B2", "label": "cap",
                 "formula": "Assumptions!B2 ^ 2 * 1e9"},
            ],
            "scenarios": [
                {"name": "bear", "overrides": {"ratio": 30.0}},
            ],
        }

    def test_live_formula_in_workbook(self):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(build_workbook(self._spec())))
        assert wb["ModelLive"]["B2"].value == "=Assumptions!B2 ^ 2 * 1e9"
        assert wb["Assumptions"]["B2"].value == 58.0
        assert wb["Assumptions"]["D2"].value == "derived"
        assert wb["Scenarios"]["A2"].value == "bear"

    def test_formula_listing_sheet(self):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(build_workbook(self._spec())))
        row = [c.value for c in wb["Model"][2]]
        assert row[0] == "cap" and row[2].startswith("=")

    def test_invalid_cell_skipped_not_crash(self):
        pytest.importorskip("openpyxl")
        spec = self._spec()
        spec["model"].append({"cell": "NOTACELL", "label": "junk",
                              "formula": "1+1"})
        wb_bytes = build_workbook(spec)  # must not raise
        assert wb_bytes[:4] == b"PK\x03\x04"

    def test_store_workbook_marks_liveness(self, store):
        out = store_workbook(self._spec(), store)
        meta = store.get_meta(out["workbook"].sha256)
        assert meta["meta"]["live_formulas"] == out["live_formulas"]
        assert out["workbook"].kind == ("xlsx" if out["live_formulas"] else "csv")

    def test_csv_fallback_lists_formulas(self):
        text = build_workbook_csv_fallback(self._spec())
        assert "Assumptions!B2 ^ 2 * 1e9" in text
        assert "58.0" in text


class TestDomainGenerality:
    """Nothing above may assume finance or sports."""

    def test_protein_chart_round_trips(self, store):
        out = store_chart(
            chart_spec("folding energy", {"E": [1.2, 0.9, 0.4, 0.5]},
                       x=[1, 2, 3, 4], x_label="residue", y_label="kcal/mol",
                       code="E = fold(seq)"),
            store,
        )
        assert store.get_json(out["spec"].sha256)["y_label"] == "kcal/mol"

    def test_supply_chain_workbook(self, tmp_path):
        pytest.importorskip("openpyxl")
        spec = {
            "assumptions": [{"name": "lead_time_days", "value": 14}],
            "data": {"Legs": {"columns": ["from", "to", "days"],
                              "rows": [["port", "dc", 12]]}},
            "model": [{"cell": "B2", "label": "total",
                       "formula": "SUM(Legs!C2:C100) + Assumptions!B2"}],
            "scenarios": [{"name": "strike", "overrides": {"lead_time_days": 40}}],
        }
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(build_workbook(spec)))
        assert "Legs" in wb.sheetnames

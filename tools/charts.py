"""S2 — Charts and spreadsheets as first-class, inspectable artifacts.

Two emitters, both domain-general (series in, artifact out):

1. Charts. matplotlib when importable; otherwise a dependency-free SVG
   line-chart renderer. Either way the artifact carries its own provenance:
   `chart_spec()` returns the series, labels, and the code_sha256 that
   produced them, so the picture and the math that generated it travel
   together. A chart you cannot regenerate from its spec is prose.

2. Spreadsheets with LIVE formulas. openpyxl builds a four-sheet workbook —
   Assumptions (every input, with source notes), Data (raw pulls with
   provenance columns), Model (formula cells referencing Assumptions/Data,
   so the owner can audit and torture the math in Excel), Scenarios
   (parameter grid feeding the same formula chain). A dead CSV of results is
   prose in disguise; an auditable formula chain IS "the math behind it".
   Falls back to a formula-listing CSV when openpyxl is absent — never
   silently: the fallback artifact is marked kind=csv with
   meta["live_formulas"]=False.
"""
from __future__ import annotations

from typing import Any, Optional

from tools.artifacts import ArtifactStore, sha256_bytes, default_store

# --------------------------------------------------------------------------
# Charts
# --------------------------------------------------------------------------


def chart_spec(
    title: str,
    series: dict[str, list[float]],
    x: Optional[list[float]] = None,
    x_label: str = "",
    y_label: str = "",
    code: str = "",
    notes: str = "",
) -> dict:
    """The provenance payload stored alongside every chart artifact.

    `series` maps label → y-values; `x` (optional) is the shared x-axis.
    `code` is the sandbox code that computed the series — sealed by hash so
    the chart is regenerable, not decorative.
    """
    lengths = {len(v) for v in series.values()}
    if x is not None and len(x) not in lengths and lengths:
        raise ValueError(
            f"x length {len(x)} does not match series lengths {sorted(lengths)}"
        )
    if len(lengths) > 1:
        raise ValueError(f"series lengths differ: {sorted(lengths)}")
    return {
        "title": title,
        "x": x,
        "x_label": x_label,
        "y_label": y_label,
        "series": series,
        "code_sha256": sha256_bytes(code.encode("utf-8")) if code else "",
        "code": code,
        "notes": notes,
    }


def render_svg(spec: dict, width: int = 720, height: int = 420) -> str:
    """Dependency-free line chart. Deterministic: same spec, same bytes."""
    series = spec["series"]
    x = spec.get("x")
    n = len(next(iter(series.values()))) if series else 0
    if n == 0:
        raise ValueError("no series to plot")

    pad_l, pad_r, pad_t, pad_b = 60, 20, 30, 44
    plot_w, plot_h = width - pad_l - pad_r, height - pad_t - pad_b

    xs = x if x is not None else list(range(n))
    all_y = [v for vals in series.values() for v in vals]
    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(all_y), max(all_y)
    if y_max == y_min:
        y_max, y_min = y_min + 1, y_min - 1

    def sx(v):  # noqa: E731
        return pad_l + (v - x_min) / (x_max - x_min or 1) * plot_w

    def sy(v):  # noqa: E731
        return pad_t + (1 - (v - y_min) / (y_max - y_min)) * plot_h

    colors = ["#1f77b4", "#d62728", "#2ca02c", "#9467bd", "#ff7f0e"]
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}">',
        f'<rect width="{width}" height="{height}" fill="white"/>',
        f'<text x="{pad_l}" y="20" font-size="15" font-family="sans-serif" '
        f'font-weight="bold">{_esc(spec["title"])}</text>',
    ]
    # y gridlines (5)
    for i in range(5):
        v = y_min + (y_max - y_min) * i / 4
        yy = sy(v)
        parts.append(
            f'<line x1="{pad_l}" y1="{yy:.1f}" x2="{width - pad_r}" y2="{yy:.1f}" '
            f'stroke="#e0e0e0"/>'
        )
        parts.append(
            f'<text x="{pad_l - 6}" y="{yy + 4:.1f}" font-size="10" '
            f'font-family="sans-serif" text-anchor="end">{_fmt(v)}</text>'
        )
    # x ticks (6)
    for i in range(6):
        v = x_min + (x_max - x_min) * i / 5
        parts.append(
            f'<text x="{sx(v):.1f}" y="{height - pad_b + 16}" font-size="10" '
            f'font-family="sans-serif" text-anchor="middle">{_fmt(v)}</text>'
        )
    # axes
    parts.append(
        f'<line x1="{pad_l}" y1="{pad_t}" x2="{pad_l}" y2="{height - pad_b}" stroke="black"/>'
    )
    parts.append(
        f'<line x1="{pad_l}" y1="{height - pad_b}" x2="{width - pad_r}" '
        f'y2="{height - pad_b}" stroke="black"/>'
    )
    if spec.get("x_label"):
        parts.append(
            f'<text x="{pad_l + plot_w / 2}" y="{height - 8}" font-size="11" '
            f'font-family="sans-serif" text-anchor="middle">{_esc(spec["x_label"])}</text>'
        )
    if spec.get("y_label"):
        parts.append(
            f'<text x="14" y="{pad_t + plot_h / 2}" font-size="11" '
            f'font-family="sans-serif" text-anchor="middle" '
            f'transform="rotate(-90 14 {pad_t + plot_h / 2})">{_esc(spec["y_label"])}</text>'
        )
    for i, (label, vals) in enumerate(series.items()):
        color = colors[i % len(colors)]
        pts = " ".join(f"{sx(xs[j]):.1f},{sy(v):.1f}" for j, v in enumerate(vals))
        parts.append(
            f'<polyline fill="none" stroke="{color}" stroke-width="2" points="{pts}"/>'
        )
        parts.append(
            f'<text x="{width - pad_r - 4}" y="{pad_t + 14 + i * 15}" font-size="11" '
            f'font-family="sans-serif" fill="{color}" text-anchor="end">{_esc(label)}</text>'
        )
    parts.append("</svg>")
    return "\n".join(parts)


def _esc(s: str) -> str:
    return (
        str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )


def _fmt(v: float) -> str:
    if abs(v) >= 1000 or (abs(v) < 0.01 and v != 0):
        return f"{v:.2e}"
    return f"{v:,.4g}"


def store_chart(
    spec: dict,
    store: Optional[ArtifactStore] = None,
    *,
    prefer_matplotlib: bool = True,
) -> dict:
    """Render the spec and store (png|svg, spec json) as linked artifacts.

    Returns {"chart": ArtifactRef, "spec": ArtifactRef, "renderer": str}.
    The chart artifact's data_refs points at the spec; the spec carries the
    code hash. Citing both = citing the picture AND its regeneration recipe.
    """
    store = store or default_store()
    renderer = "svg"
    kind = "svg"
    data: Optional[bytes] = None
    if prefer_matplotlib:
        try:
            import io

            import matplotlib

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            fig, ax = plt.subplots(figsize=(7.2, 4.2))
            for label, vals in spec["series"].items():
                xs = spec.get("x") or list(range(len(vals)))
                ax.plot(xs, vals, label=label)
            ax.set_title(spec["title"])
            ax.set_xlabel(spec.get("x_label", ""))
            ax.set_ylabel(spec.get("y_label", ""))
            ax.legend()
            ax.grid(alpha=0.3)
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
            plt.close(fig)
            data = buf.getvalue()
            renderer, kind = "matplotlib", "png"
        except ImportError:
            data = None
    if data is None:
        data = render_svg(spec).encode("utf-8")

    code = spec.pop("code", "")  # code lives in the spec artifact, not the image
    spec_ref = store.put_json(
        {**spec, "code": code},
        name=f"{spec['title'][:40]} spec",
        meta={"renderer": renderer},
    )
    chart_ref = store.put(
        data,
        kind,
        name=f"{spec['title'][:40]}",
        code_sha256=spec.get("code_sha256", ""),
        data_refs=[spec_ref.sha256],
    )
    return {"chart": chart_ref, "spec": spec_ref, "renderer": renderer}


# --------------------------------------------------------------------------
# Spreadsheets with live formulas
# --------------------------------------------------------------------------

WORKBOOK_SHEETS = ("Assumptions", "Data", "Model", "Scenarios")


def build_workbook(spec: dict) -> bytes:
    """Build the standard 4-sheet xlsx from a workbook spec.

    Spec shape (all domain-general — keys are column names, not finance):
      assumptions: [{name, value, unit, source, note}]
      data:        {sheet_name: {columns: [...], rows: [[...], ...],
                                 provenance: [{column, source, fetched_at}]}}
      model:       [{cell, formula, label}]   # formulas are LIVE Excel cells
      scenarios:   [{name, overrides: {assumption_name: value}}]

    The Model sheet's formulas reference Assumptions!B<row> and Data ranges,
    so changing an assumption recomputes the model inside Excel. That is the
    auditability requirement: the owner can torture every input.
    """
    try:
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for live-formula workbooks; "
            "pip install openpyxl"
        ) from exc

    wb = Workbook()
    bold = Font(bold=True)

    # Assumptions ---------------------------------------------------------
    ws = wb.active
    ws.title = "Assumptions"
    ws.append(["name", "value", "unit", "source", "note"])
    for c in ws[1]:
        c.font = bold
    assumption_row: dict[str, int] = {}
    for a in spec.get("assumptions", []):
        ws.append([a["name"], a["value"], a.get("unit", ""),
                   a.get("source", ""), a.get("note", "")])
        assumption_row[a["name"]] = ws.max_row
    for col, w in zip("ABCDE", (28, 14, 10, 34, 40)):
        ws.column_dimensions[col].width = w

    # Data ----------------------------------------------------------------
    for sheet_name, table in spec.get("data", {}).items():
        ws = wb.create_sheet(title=sheet_name[:31])
        cols = table["columns"]
        ws.append(cols)
        for c in ws[1]:
            c.font = bold
        for row in table.get("rows", []):
            ws.append(list(row))
        for p in table.get("provenance", []):
            col_idx = cols.index(p["column"]) + 1 if p["column"] in cols else 1
            note = f"source: {p.get('source', '')} fetched: {p.get('fetched_at', '')}"
            ws.cell(row=1, column=col_idx).comment = _mk_comment(note)
        ws.freeze_panes = "A2"

    # Model — the listing sheet documents each formula; ModelLive holds the
    # same formulas as live Excel cells so the workbook actually computes.
    ws = wb.create_sheet("Model")
    for i, h in enumerate(["label", "cell", "formula"]):
        ws.cell(row=1, column=i + 1, value=h).font = bold
    for r, m in enumerate(spec.get("model", []), start=2):
        ws.cell(row=r, column=1, value=m.get("label", ""))
        ws.cell(row=r, column=2, value=m["cell"])
        ws.cell(row=r, column=3, value="=" + m["formula"].lstrip("="))
    ws_live = wb.create_sheet("ModelLive")
    for m in spec.get("model", []):
        if _valid_cell(m["cell"]):
            ws_live[m["cell"]] = "=" + m["formula"].lstrip("=")
            header = ws_live.cell(row=1, column=_col_index(m["cell"]))
            if not header.value:
                header.value = m.get("label", m["cell"])

    # Scenarios ------------------------------------------------------------
    ws = wb.create_sheet("Scenarios")
    ws.append(["scenario", "assumption", "value"])
    for c in ws[1]:
        c.font = bold
    for sc in spec.get("scenarios", []):
        for k, v in sc["overrides"].items():
            ws.append([sc["name"], k, v])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mk_comment(text: str):
    from openpyxl.comments import Comment

    return Comment(text, "Callisto")


def _col_index(cell: str) -> int:
    letters = "".join(ch for ch in cell if ch.isalpha()).upper()
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _valid_cell(cell: str) -> bool:
    import re

    return bool(re.fullmatch(r"[A-Za-z]{1,3}[1-9][0-9]*", cell))


def build_workbook_csv_fallback(spec: dict) -> str:
    """Human-readable formula listing when openpyxl is missing. Explicitly
    NOT a live workbook — meta marks it so nothing mistakes it for one."""
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["sheet", "item", "value_or_formula"])
    for a in spec.get("assumptions", []):
        w.writerow(["Assumptions", a["name"], a["value"]])
        if a.get("source"):
            w.writerow(["Assumptions", f"{a['name']} source", a["source"]])
    for sheet_name, table in spec.get("data", {}).items():
        w.writerow([sheet_name, "columns", ",".join(table["columns"])])
        for row in table.get("rows", []):
            w.writerow([sheet_name, "row", ",".join(str(v) for v in row)])
    for m in spec.get("model", []):
        w.writerow(["Model (formulas, not live)", m.get("label", ""), m["formula"]])
    for sc in spec.get("scenarios", []):
        for k, v in sc["overrides"].items():
            w.writerow(["Scenarios", f"{sc['name']}:{k}", v])
    return buf.getvalue()


def store_workbook(
    spec: dict,
    store: Optional[ArtifactStore] = None,
    name: str = "model_workbook",
) -> dict:
    """Emit the workbook artifact. Live xlsx when openpyxl exists; otherwise
    an explicitly-degraded CSV listing. Returns refs + a `live_formulas`
    flag callers must surface in the conclusion."""
    store = store or default_store()
    code = spec.get("code", "")
    try:
        data = build_workbook(spec)
        ref = store.put(
            data, "xlsx", name=name,
            code_sha256=sha256_bytes(code.encode("utf-8")) if code else "",
            meta={"live_formulas": True, "sheets": WORKBOOK_SHEETS},
        )
        return {"workbook": ref, "live_formulas": True}
    except RuntimeError:
        csv_text = build_workbook_csv_fallback(spec)
        ref = store.put_text(
            csv_text, "csv", name=f"{name}_fallback",
            code_sha256=sha256_bytes(code.encode("utf-8")) if code else "",
            meta={"live_formulas": False,
                  "degraded": "openpyxl unavailable; install it for a live workbook"},
        )
        return {"workbook": ref, "live_formulas": False}
